#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_SOURCE = (
    "/Users/lixin/Documents/大宜宾/其他资料/志愿填报/"
    "23、四川-2026志愿填报资料【先保存才能用】/四川-2026计划-专家版数据（23-26）.xlsx"
)
DEFAULT_NDJSON = "data/gaokao/sichuan-2026-major-plans.ndjson"
DEFAULT_MARKDOWN = "docs/gaokao-sichuan-2026-official-major-plan-data.md"


FIELD_MAP = {
    "ID": "source_row_id",
    "年份": "year",
    "生源地": "province",
    "科类": "subject_type",
    "批次": "batch",
    "计划类别": "plan_type",
    "院校代码": "school_code",
    "院校名称": "school_name",
    "院校专业组代码": "school_major_group_code",
    "专业组代码": "group_code",
    "专业组名称": "group_name",
    "专业代码": "major_code",
    "专业全称": "major_full_name",
    "专业名称": "major_name",
    "专业备注": "major_note",
    "其他备注": "other_note",
    "选科要求": "subject_requirement",
    "专业层次": "degree_level",
    "计划人数": "plan_count",
    "学费": "tuition",
    "学制": "duration",
    "专业组计划人数": "group_plan_count",
    "组内专业数": "group_major_count",
    "专业组干净度": "group_cleanliness",
    "门类": "major_category",
    "专业类": "major_class",
    "26年预估位次": "estimated_rank_2026",
    "是否新增": "is_new",
    "专业组录取人数1": "group_admit_count_2025",
    "专业组最低分1": "group_score_2025",
    "专业组最低位次1": "group_rank_2025",
    "录取人数1": "admit_count_2025",
    "最低分1": "score_2025",
    "最低位次1": "rank_2025",
    "平均分1": "average_score_2025",
    "平均位次1": "average_rank_2025",
    "录取人数2": "admit_count_2024",
    "最低分2": "score_2024",
    "最低位次2": "rank_2024",
    "平均分2": "average_score_2024",
    "平均位次2": "average_rank_2024",
    "录取人数3": "admit_count_2023",
    "最低分3": "score_2023",
    "最低位次3": "rank_2023",
    "平均分3": "average_score_2023",
    "平均位次3": "average_rank_2023",
    "所在省": "school_province",
    "城市": "school_city",
    "城市水平标签": "city_level_label",
    "院校标签": "school_tags",
    "院校水平": "school_level",
    "类型": "school_type",
    "公私性质": "ownership",
    "院校排名": "school_rank",
    "保研率": "postgraduate_rate",
    "转专业情况": "transfer_policy",
    "录取规则": "admission_rule",
    "招生章程": "charter_url",
    "软科评级": "major_rating",
    "软科排名": "major_ranking",
    "学科评估": "discipline_eval",
    "专业水平": "major_level",
}

INT_FIELDS = {
    "year",
    "plan_count",
    "tuition",
    "group_plan_count",
    "group_major_count",
    "estimated_rank_2026",
    "group_admit_count_2025",
    "group_score_2025",
    "group_rank_2025",
    "admit_count_2025",
    "score_2025",
    "rank_2025",
    "average_score_2025",
    "average_rank_2025",
    "admit_count_2024",
    "score_2024",
    "rank_2024",
    "average_score_2024",
    "average_rank_2024",
    "admit_count_2023",
    "score_2023",
    "rank_2023",
    "average_score_2023",
    "average_rank_2023",
    "school_rank",
    "major_ranking",
}

FLOAT_FIELDS = {"group_cleanliness", "postgraduate_rate"}


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_int(value: Any) -> int | None:
    text = clean_text(value)
    if not text:
        return None
    text = text.replace(",", "").replace("，", "").replace("人", "").replace("元", "")
    try:
        return int(float(text))
    except ValueError:
        return None


def to_float(value: Any) -> float | None:
    text = clean_text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_subject(value: Any) -> str:
    text = clean_text(value) or ""
    if "历史" in text or "文科" in text:
        return "历史类"
    if "物理" in text or "理科" in text:
        return "物理类"
    return text or "未标注"


def normalize_bool(value: Any) -> bool:
    text = clean_text(value) or ""
    return text not in {"", "否", "无", "0", "False", "false", "None"}


def normalize_row(headers: list[str], values: tuple[Any, ...], source_file: str) -> dict[str, Any] | None:
    raw = {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}
    source_row_id = clean_text(raw.get("ID"))
    school_name = clean_text(raw.get("院校名称"))
    major_name = clean_text(raw.get("专业名称"))

    if not source_row_id or not school_name or not major_name:
        return None

    item: dict[str, Any] = {
        "source_id": f"sichuan2026-major-plan:{source_row_id}",
        "source_file": source_file,
    }

    for source_key, target_key in FIELD_MAP.items():
        value = raw.get(source_key)
        if target_key == "source_row_id":
            continue
        if target_key == "subject_type":
            item[target_key] = normalize_subject(value)
        elif target_key == "is_new":
            item[target_key] = normalize_bool(value)
        elif target_key in INT_FIELDS:
            item[target_key] = to_int(value)
        elif target_key in FLOAT_FIELDS:
            item[target_key] = to_float(value)
        else:
            item[target_key] = clean_text(value)

    if item.get("province") != "四川" or item.get("year") != 2026:
        return None

    return item


def md_escape(value: Any) -> str:
    text = "" if value is None else str(value)
    return text.replace("|", "\\|").replace("\n", " ")


def render_markdown(
    source: Path,
    output: Path,
    rows: list[dict[str, Any]],
    headers: list[str],
) -> None:
    subject_counts = Counter(row["subject_type"] for row in rows)
    batch_counts = Counter(row.get("batch") for row in rows)
    schools = {row.get("school_name") for row in rows if row.get("school_name")}
    groups = {
        (
            row.get("school_name"),
            row.get("group_code"),
            row.get("subject_type"),
            row.get("batch"),
        )
        for row in rows
    }
    majors = {row.get("major_name") for row in rows if row.get("major_name")}
    category_counts = Counter(row.get("major_category") or "未标注" for row in rows)

    by_group: dict[tuple[Any, ...], list[str]] = defaultdict(list)
    computer_samples: list[dict[str, Any]] = []
    for row in rows:
        key = (
            row.get("school_name"),
            row.get("group_code"),
            row.get("subject_type"),
            row.get("batch"),
        )
        if len(by_group[key]) < 8:
            by_group[key].append(row.get("major_name") or "")
        text = " ".join(
            str(row.get(field) or "")
            for field in ("school_name", "major_name", "major_note", "major_category", "major_class")
        )
        if "计算机" in text and len(computer_samples) < 10:
            computer_samples.append(row)

    group_examples = [
        (key, values)
        for key, values in by_group.items()
        if key[0] and len([value for value in values if value]) >= 4
    ][:10]

    lines = [
        "# 四川 2026 官方招生计划专业组数据提取说明",
        "",
        "本文件是从本地官方志愿填报资料中抽取出的可审计数据说明。全量结构化数据已写入 `data/gaokao/sichuan-2026-major-plans.ndjson`，并可导入 `gaokao_major_plans` 数据表。",
        "",
        "## 来源文件",
        "",
        f"- `{source}`",
        "",
        "## 提取结果",
        "",
        f"- 2026 四川招生计划专业记录：{len(rows):,} 条",
        f"- 院校数量：{len(schools):,} 所",
        f"- 院校专业组数量：{len(groups):,} 个",
        f"- 专业名称数量：{len(majors):,} 个",
        f"- 科类分布：{', '.join(f'{k} {v:,}' for k, v in subject_counts.items())}",
        "",
        "## 批次分布 Top 12",
        "",
        "| 批次 | 记录数 |",
        "|---|---:|",
    ]
    lines.extend(f"| {md_escape(key)} | {value:,} |" for key, value in batch_counts.most_common(12))

    lines.extend([
        "",
        "## 专业门类分布 Top 12",
        "",
        "| 门类 | 记录数 |",
        "|---|---:|",
    ])
    lines.extend(f"| {md_escape(key)} | {value:,} |" for key, value in category_counts.most_common(12))

    lines.extend([
        "",
        "## 已保留的核心字段",
        "",
        "| 原字段 | 用途 |",
        "|---|---|",
    ])
    for header in headers:
        if header in FIELD_MAP and header != "ID":
            lines.append(f"| {md_escape(header)} | `{FIELD_MAP[header]}` |")

    lines.extend([
        "",
        "## 计算机方向样例",
        "",
        "| 科类 | 批次 | 院校 | 专业组 | 专业 | 备注 | 计划 | 2026预估位次 | 2025专业位次 | 学费 |",
        "|---|---|---|---|---|---|---:|---:|---:|---:|",
    ])
    for row in computer_samples:
        lines.append(
            "| "
            + " | ".join(
                md_escape(row.get(field))
                for field in (
                    "subject_type",
                    "batch",
                    "school_name",
                    "group_code",
                    "major_name",
                    "major_note",
                    "plan_count",
                    "estimated_rank_2026",
                    "rank_2025",
                    "tuition",
                )
            )
            + " |"
        )

    lines.extend([
        "",
        "## 院校专业组样例",
        "",
        "| 院校 | 专业组 | 科类 | 批次 | 组内专业样例 |",
        "|---|---|---|---|---|",
    ])
    for key, values in group_examples:
        lines.append(
            "| "
            + " | ".join(md_escape(value) for value in (*key, "、".join(values)))
            + " |"
        )

    lines.extend([
        "",
        "## 对用户端推荐的直接价值",
        "",
        "- 可以把现有“院校专业组”推荐补成“组内建议专业”。",
        "- 可以识别组内专业是否集中，避免把混杂专业组误说成某个具体专业稳妥。",
        "- 可以展示学费、学制、计划人数、选科要求、往年专业最低位次、院校标签和专业评级。",
        "- 可以对用户偏好专业做排序，而不是只按学校名和专业组历史位次排序。",
        "",
        "## 注意",
        "",
        "- 本文件不展开 5 万多行明细，避免 Markdown 变成不可读的大表。",
        "- 全量明细以 NDJSON 和数据库为准；本文件用于人工审计、字段核对和产品方案沟通。",
        "",
    ])

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=DEFAULT_SOURCE)
    parser.add_argument("--ndjson", default=DEFAULT_NDJSON)
    parser.add_argument("--markdown", default=DEFAULT_MARKDOWN)
    args = parser.parse_args()

    source = Path(args.source)
    ndjson = Path(args.ndjson)
    markdown = Path(args.markdown)

    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb.active
    headers = [clean_text(value) or "" for value in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    rows: list[dict[str, Any]] = []

    ndjson.parent.mkdir(parents=True, exist_ok=True)
    with ndjson.open("w", encoding="utf-8") as handle:
        for values in ws.iter_rows(min_row=3, values_only=True):
            item = normalize_row(headers, values, str(source))
            if not item:
                continue
            rows.append(item)
            handle.write(json.dumps(item, ensure_ascii=False, separators=(",", ":")) + "\n")

    wb.close()
    render_markdown(source, markdown, rows, headers)

    print(json.dumps({
        "source": str(source),
        "ndjson": str(ndjson),
        "markdown": str(markdown),
        "rows": len(rows),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
